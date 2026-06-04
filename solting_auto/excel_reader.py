"""엑셀 읽기. 고정 양식(주민번호/이름/전화번호) 기준. (PRD FR-1.3)"""

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


@dataclass
class Record:
    row_no: int          # 엑셀 행 번호(헤더=1 기준, 데이터는 2부터)
    jumin: str
    name: str
    phone: str
    raw: dict            # 원본 셀 값


def read_records(xlsx_path: str, columns: dict) -> list:
    """엑셀에서 레코드 목록을 읽는다.

    columns 예: {"jumin": "주민번호", "name": "이름", "phone": "전화번호"}
    """
    path = Path(xlsx_path)
    if not path.exists():
        raise FileNotFoundError(f"엑셀 파일이 없습니다: {xlsx_path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    rows = ws.iter_rows(values_only=True)
    try:
        header = [(_s(c)) for c in next(rows)]
    except StopIteration:
        wb.close()
        raise ValueError("엑셀이 비어 있습니다.")

    # 컬럼 인덱스 매핑
    idx = {}
    for key, colname in columns.items():
        if colname not in header:
            wb.close()
            raise ValueError(
                f"필수 컬럼 '{colname}' 이(가) 엑셀 헤더에 없습니다. "
                f"엑셀 헤더: {header}"
            )
        idx[key] = header.index(colname)

    records = []
    for i, row in enumerate(rows, start=2):
        # 완전 빈 행은 건너뜀
        if row is None or all(c is None or _s(c) == "" for c in row):
            continue
        raw = {columns[k]: _cell(row, idx[k]) for k in columns}
        records.append(
            Record(
                row_no=i,
                jumin=_cell(row, idx["jumin"]),
                name=_cell(row, idx["name"]),
                phone=_cell(row, idx["phone"]),
                raw=raw,
            )
        )
    wb.close()
    return records


def _cell(row, i):
    if i >= len(row):
        return ""
    return _s(row[i])


def _s(v):
    if v is None:
        return ""
    return str(v).strip()
